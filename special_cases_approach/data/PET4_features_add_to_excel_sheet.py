#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul 29 16:06:34 2022

@author: kouame
"""

import openpyxl as xl

filename1 ="PET4_plus.xlsx"; wb1 = xl.load_workbook(filename1); pet_4_plus = wb1.worksheets[0]
filename2 ="classical_features.xlsx"; wb2 = xl.load_workbook(filename2); features = wb2.worksheets[0] 

for row_pet4 in pet_4_plus.iter_rows():
    for row_features in features.iter_rows():
        if str(row_pet4[0].value) == row_features[0].value:
            row_pet4[5].value = (row_features[1].value)*0.001
            row_pet4[10].value = row_features[2].value
            row_pet4[15].value = row_features[3].value
            row_pet4[18].value = row_features[4].value
wb1.save(filename1) ###### A verifier avant d'exécuter le programme############
            