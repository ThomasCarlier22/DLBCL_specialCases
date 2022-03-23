# -*- coding: utf-8 -*-
"""
Created on Thu Mar  3 13:22:05 2022

@author: Fero
"""

import openpyxl as xl; 
import os
import shutil as st
import numpy as np

def rm (filename): #rm supprime le dossier filename  s'il existe
    if os.path.exists(filename):
       os.remove(filename)
    return 


def rmd (filename): #rm supprime le dossier filename  s'il existe
    if os.path.exists(filename):
       st.rmtree(filename)
    return  


def SUV_max_ref(SUVmax_ref1,SUVmax_ref2,suv_max_adjudication):
    ref=0
    methode = ""
    if isinstance(SUVmax_ref1,(int,float)) and isinstance(SUVmax_ref2,(int,float)) and  SUVmax_ref1!=SUVmax_ref2 and isinstance(suv_max_adjudication,(int,float)) :
        ref=suv_max_adjudication
        methode = "ADJ"    
    elif isinstance(SUVmax_ref1,(int,float)) and isinstance(SUVmax_ref2,(int,float)) and SUVmax_ref1!=SUVmax_ref2 and ((not(isinstance(suv_max_adjudication,(int,float)))) or suv_max_adjudication==0) :
        ref=(SUVmax_ref1+SUVmax_ref2)/2;  methode = "Moy_1_2" 
        
    elif SUVmax_ref1==SUVmax_ref2 :
        ref=SUVmax_ref1;  methode = "Egalité_1_2"
        
    else:
        ref= "erreur";  methode="aucune"
        
    return [ref,methode]


#Suppression de dossiers existantse et création de nouveaux fichiers
rmd("PET");   rmd("analyse_reviewers");   rmd("accord_Deauville_Rev_1_2");   rmd("desaccord_Deauville_rev_1_2") ; rmd("cas_spéciaux");   os.system('mkdir PET') ;  os.system('mkdir analyse_reviewers'); os.system('mkdir accord_Deauville_Rev_1_2'); os.system("mkdir desaccord_Deauville_rev_1_2");   os.system("mkdir cas_spéciaux")
filename ="GAINED.xlsx"; wb = xl.load_workbook(filename); gained = wb.worksheets[0] 
filename1 ="PET/PET0.xlsx"; filename2 ="PET/PET2.xlsx"; filename3 ="PET/PET4.xlsx"; filename4 ="analyse_reviewers/special_REV1_TEP2.xlsx"; filename5 ="analyse_reviewers/special_REV1_TEP4.xlsx"; filename6 ="analyse_reviewers/special_REV2_TEP2.xlsx"; 
filename7 ="analyse_reviewers/special_REV2_TEP4.xlsx"; filename8 ="desaccord_Deauville_rev_1_2/desaccord_TEP2.xlsx"; filename9 ="desaccord_Deauville_rev_1_2/desaccord_TEP4.xlsx"; filename10 ="accord_Deauville_Rev_1_2/accord_TEP2.xlsx"; filename11 = "accord_Deauville_Rev_1_2/accord_TEP4.xlsx"
filename12="cas_spéciaux/PET2_moins.xlsx"; filename13="cas_spéciaux/PET2_plus.xlsx"; filename14="cas_spéciaux/PET4_plus.xlsx"; filename15="cas_spéciaux/total_cas_spéciaux.xlsx"; filename16="cas_spéciaux/PET2+PET4+.xlsx"; filename17="cas_spéciaux/PET2-PET4+.xlsx"



#Commandes pour ouvrir les différentes feuilles de calcul 
workbook1 = xl. Workbook(); workbook1.save(filename1); workbook2 = xl. Workbook(); workbook2.save(filename2); workbook3 = xl. Workbook(); workbook3.save(filename3); workbook4 = xl. Workbook(); workbook4.save(filename4); workbook5 = xl. Workbook(); workbook5.save(filename5); workbook6 = xl. Workbook(); workbook6.save(filename6); workbook7 = xl. Workbook(); workbook7.save(filename7); 
workbook8 = xl. Workbook(); workbook8.save(filename8); workbook9 = xl. Workbook(); workbook9.save(filename9); workbook10 = xl. Workbook(); workbook10.save(filename10); workbook11 = xl. Workbook(); workbook11.save(filename11); workbook12 = xl. Workbook(); workbook12.save(filename12); workbook13 = xl. Workbook(); workbook13.save(filename13); workbook14 = xl. Workbook(); workbook14.save(filename14); workbook15 = xl. Workbook(); workbook15.save(filename15)
workbook16 = xl. Workbook(); workbook16.save(filename16); workbook17 = xl. Workbook(); workbook17.save(filename17)
wb1 = xl.load_workbook(filename1);PET0 = wb1.active ;       wb2 = xl.load_workbook(filename2); PET2 = wb2.active  ;      wb3 = xl.load_workbook(filename3); PET4 = wb3.active;     wb4 = xl.load_workbook(filename4);REV1_PET2 = wb4.active;    wb5 = xl.load_workbook(filename5); REV1_PET4 = wb5.active;     wb6 = xl.load_workbook(filename6); REV2_PET2 = wb6.active;      wb7 = xl.load_workbook(filename7); REV2_PET4 = wb7.active;   
wb8 = xl.load_workbook(filename8); des_PET2 = wb8.active;     wb9 = xl.load_workbook(filename9); des_PET4 = wb9.active; wb10 = xl.load_workbook(filename10); accord_PET2 = wb10.active;    wb11 = xl.load_workbook(filename11); accord_PET4 = wb11.active; wb12 = xl.load_workbook(filename12); PET2_moins = wb12.active; wb13 = xl.load_workbook(filename13); PET2_plus = wb13.active; wb14 = xl.load_workbook(filename14); PET4_plus = wb14.active   
wb15 = xl.load_workbook(filename15); total_cas_spéciaux = wb15.active;   wb16 = xl.load_workbook(filename16); PET_2plus_4plus = wb16.active;   wb17 = xl.load_workbook(filename17); PET_2moins_4plus = wb17.active;



#Commandes pour ajouter le titre des colonnes à chaque feuille
l1=["Dif_SUV_max_REV_1_2_(%)", "SUV_max_ref", "methode_SUV_max_ref", "SUV_max_extrait", "condition vérifiée"]
l=[gained.cell(row = 1, column = j).value for j in range(1,67) ]
l.extend(l1)
PET0.append(l); PET2.append(l); PET4.append(l); REV1_PET2.append(l);    REV1_PET4.append(l);   REV2_PET2.append(l);   REV2_PET4.append(l) ; des_PET2.append(l);    des_PET4.append(l);  accord_PET2.append(l);   accord_PET4.append(l); PET2_moins.append(l); PET2_plus.append(l); PET4_plus.append(l); total_cas_spéciaux.append(l)




#Données fournis par Gauthier
patients_extraits=np.load("donées list_suv_gautier/list_patients.npy")
SUV_max_extrait =np.load("donées list_suv_gautier/SUV_max.npy")




#Création des listes des TEP
for row in gained.iter_rows():
    PET=row[1].value
    l=[row[j].value for j in range(len(row))]    
    if PET=='PET0':       
        PET0.append(l)
    elif PET=='PET2':
        PET2.append(l)
    elif PET=='PET4':
        PET4.append(l) 
        
        
        

#Ajout des SUV_max extraites par Gauthier
mr = PET0.max_row

for i in range(2,mr+1):
    for j in range(len(patients_extraits)):
        if PET0.cell(i,1).value == int(patients_extraits[j]):
            PET0.cell(i,70).value = SUV_max_extrait[j]
 
for i in range(2,PET0.max_row+1):
    PET0.cell(row=i,column=68).value=SUV_max_ref(PET0.cell(row=i,column=42).value, PET0.cell(row=i,column=43).value, PET0.cell(row=i,column=29).value)[0]
    PET0.cell(row=i,column=69).value=SUV_max_ref(PET0.cell(row=i,column=42).value, PET0.cell(row=i,column=43).value, PET0.cell(row=i,column=29).value)[1]
    if isinstance(PET0.cell(row=i,column=42).value,(float,int,str)) and isinstance(PET0.cell(row=i,column=43).value,(float,int,str)):
        PET0.cell(row=i,column=67).value = 100*abs((PET0.cell(row=i,column=42).value-PET0.cell(row=i,column=43).value)/max(PET0.cell(row=i,column=42).value,PET0.cell(row=i,column=43).value) )
    if isinstance(PET2.cell(row=i,column=42).value,(float,int,str)) and isinstance(PET2.cell(row=i,column=43).value,(float,int,str)):
        PET2.cell(row=i,column=67).value = 100*abs((PET2.cell(row=i,column=42).value-PET2.cell(row=i,column=43).value)/max(PET2.cell(row=i,column=42).value,PET2.cell(row=i,column=43).value) )
    if isinstance(PET4.cell(row=i,column=42).value,(float,int,str)) and isinstance(PET4.cell(row=i,column=45).value,(float,int,str)):
        PET4.cell(row=i,column=67).value = 100*abs((PET4.cell(row=i,column=42).value-PET4.cell(row=i,column=45).value)/max(PET4.cell(row=i,column=42).value,PET4.cell(row=i,column=45).value) )
   
    
   
                   
#Vérification des conditions et classemment des cas spéciaux pour chaque observateur.
for i in range(2,PET0.max_row+1):  
        if ((isinstance(PET2.cell(row = i, column = 50).value,(int,float))) and (isinstance(PET0.cell(row = i, column = 42).value,(int,float))) and ((PET2.cell(row = i, column = 50).value)<=66 and (PET0.cell(row = i, column = 42).value<10))):
            l2=[PET2.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ]; 
            l2.append("delta_SUV(TEP2)<=66% et SUV_max(TEP0)<10")
            REV1_PET2.append(l2) # REV1_PET2.append(l1);
            
        elif(((isinstance(PET2.cell(row = i, column = 50).value,(int,float))) and (isinstance(PET2.cell(row = i, column = 42).value,(int,float)))) and ((PET2.cell(row = i, column = 50).value)>66 and (PET2.cell(row = i, column = 42).value>5))):
            l2=[PET2.cell(row =i , column = j).value for j in range(1,PET0.max_column) ];#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ]; 
            l2.append("delta_SUV(TEP2)>66% et SUV_max(TEP2)>5")
            REV1_PET2.append(l2)#REV1_PET2.append(l1)
        
        if ((isinstance(PET4.cell(row = i, column = 50).value,(int,float))) and (isinstance(PET0.cell(row = i, column = 42).value,(int,float))) and ((PET4.cell(row = i, column = 50).value)<=70 and (PET0.cell(row = i, column = 42).value<10))) :               
           l2=[PET4.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]# l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ];
           l2.append("delta_SUV(TEP4)<=70% et SUV_max(TEP0)<10")
           REV1_PET4.append(l2)# REV1_PET4.append(l1); 
            
        elif (((isinstance(PET4.cell(row = i, column = 50).value,(int,float))) and (isinstance(PET4.cell(row = i, column = 42).value,(int,float)))) and ((PET4.cell(row = i, column = 50).value)>70 and (PET4.cell(row = i, column = 42).value>5))):
            l2=[PET4.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ];
            l2.append("delta_SUV(TEP4)>70% et SUV_max(TEP4)>5")
            REV1_PET4.append(l2)#REV1_PET4.append(l1); 
            
    
        if ((isinstance(PET2.cell(row = i, column = 51).value,(int,float))) and (isinstance(PET0.cell(row = i, column = 43).value,(int,float))) and ((PET2.cell(row = i, column = 51).value)<=66 and (PET0.cell(row = i, column = 43).value<10))) :    
            l2=[PET2.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ]; 
            l2.append("delta_SUV(TEP2)<=66% et SUV_max(TEP0)<10")
            REV2_PET2.append(l2)#REV2_PET2.append(l1);  
            
        elif  (((isinstance(PET2.cell(row = i, column = 51).value,(int,float))) and (isinstance(PET2.cell(row = i, column = 43).value,(int,float)))) and ((PET2.cell(row = i, column = 51).value)>66 and (PET2.cell(row = i, column = 43).value>5))):   
            l2=[PET2.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ]; 
            l2.append("delta_SUV(TEP2)>66% et SUV_max(TEP2)>5")
            REV2_PET2.append(l2)#REV2_PET2.append(l1);  
            
          
        if ((isinstance(PET4.cell(row = i, column = 51).value,(int,float))) and (isinstance(PET0.cell(row = i, column = 43).value,(int,float))) and ((PET4.cell(row = i, column = 50).value)<=70 and (PET0.cell(row = i, column = 42).value<10))) :     
            l2=[PET4.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ];
            l2.append("delta_SUV(TEP4)<=70% et SUV_max(TEP0)<10")
            REV2_PET4.append(l2)#REV2_PET4.append(l1);
            
        elif  (((isinstance(PET4.cell(row = i, column = 51).value,(int,float))) and (isinstance(PET4.cell(row = i, column = 45).value,(int,float)))) and ((PET4.cell(row = i, column = 51).value)>70  and (PET4.cell(row = i, column = 45).value>5))):   
            l2=[PET4.cell(row =i , column = j).value for j in range(1,PET0.max_column) ]#l1=[PET0.cell(row =i , column = j).value for j in range(1,PET0.max_column+1) ];
            l2.append("delta_SUV(TEP4)>70% et SUV_max(TEP4)>5")
            REV2_PET4.append(l2)#REV2_PET4.append(l1);

    
            
   
#List des numero d'inclusion des patients dans chaque cas            
List_PET2_REV1 = [REV1_PET2.cell(j,1).value for j in range(2,REV1_PET2.max_row+1)]
List_PET2_REV2 = [REV2_PET2.cell(j,1).value for j in range(2,REV2_PET2.max_row+1)]
List_PET4_REV1 = [REV1_PET4.cell(j,1).value for j in range(2,REV1_PET4.max_row+1)]
List_PET4_REV2 = [REV2_PET4.cell(j,1).value for j in range(2,REV2_PET4.max_row+1)]




#Parcours des liste des numero d'inclusions pour détecter les accords et désaccords
for i in range(2,REV1_PET2.max_row+1):
    if REV1_PET2.cell(i,1).value not in List_PET2_REV2:
        l=[REV1_PET2.cell(i,j).value for j in range(1,PET0.max_column+1)]
        l.append("rev1")
        des_PET2.append(l)
                                                                                                                        #des_PET2.append([REV1_PET2.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
for i in range(2,REV2_PET2.max_row+1):
    if REV2_PET2.cell(i,1).value not in List_PET2_REV1:
        l=[REV2_PET2.cell(i,j).value for j in range(1,PET0.max_column+1)]
        l.append("Rev2")
        des_PET2.append(l)
                                                                                                                         #des_PET2.append([REV2_PET2.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
    else:
     accord_PET2.append([REV2_PET2.cell(i,j).value for j in range(1,PET0.max_column+1)])
                                                                                                                         #accord_PET2.append([REV2_PET2.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
             
for i in range(2,REV1_PET4.max_row+1):
    if REV1_PET4.cell(i,1).value not in List_PET4_REV2:
        l=[REV1_PET4.cell(i,j).value for j in range(1,PET0.max_column+1)]
        l.append("Rev1")
        des_PET4.append(l)
                                                                                                                        #des_PET4.append([REV1_PET4.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
for i in range(2,REV2_PET4.max_row+1):
    if REV2_PET4.cell(i,1).value not in List_PET4_REV1:
        l=[REV2_PET4.cell(i,j).value for j in range(1,PET0.max_column+1)]
        l.append("Rev2")
        des_PET4.append(l)
                                                                                                                        #des_PET4.append([REV2_PET4.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
    else:
        accord_PET4.append([REV2_PET4.cell(i,j).value for j in range(1,PET0.max_column+1)])



#Suppression des patients desaccord REV1_REV2 en TEP2 non évalués par Dauville
mr=des_PET2.max_row
for i in range(2,mr+1):
    if des_PET2.cell(i, 62).value==des_PET2.cell(i, 59).value   and   des_PET2.cell(i, 59).value!=des_PET2.cell(i, 58).value   and  (not isinstance(des_PET2.cell(i, 55).value,(int,float))):
        for j in range(1,des_PET2.max_column+1):
            des_PET2.cell(i,j).value=""
          
    if des_PET2.cell(i, 62).value==des_PET2.cell(i, 58).value   and   des_PET2.cell(i, 58).value!=des_PET2.cell(i, 59).value   and  (not isinstance(des_PET2.cell(i, 54).value,(int,float))):
        for j in range(1,des_PET2.max_column+1):
            des_PET2.cell(i,j).value=""
    
    if des_PET2.cell(i, 62).value==des_PET2.cell(i, 58).value   and   des_PET2.cell(i, 62).value==des_PET2.cell(i, 59).value   and  (not isinstance(des_PET2.cell(i, 39).value,(int,float,str))):
        for j in range(1,des_PET2.max_column+1):
            des_PET2.cell(i,j).value=""
                
 
                                                                                                             #accord_PET4.append([REV2_PET4.cell(i+1,j).value for j in range(1,PET0.max_column+1)])
#Determination des TEP4+ 
for row in accord_PET4.iter_rows():
    if row[61].value=="Positive":
        l=[row[j].value for j in range (PET0.max_column)]
        PET4_plus.append(l)        
for row in des_PET4.iter_rows():
    if row[61].value=="Positive":
        l=[row[j].value for j in range (PET0.max_column)]
        PET4_plus.append(l)



#Determination de TEP2+ et TEP2-
for row in accord_PET2.iter_rows():
    if row[61].value=="Positive":
        l=[row[j].value for j in range (PET0.max_column)]
        PET2_plus.append(l)
    elif row[61].value=="Negative":
        l=[row[j].value for j in range (PET0.max_column)]
        PET2_moins.append(l)
for row in des_PET2.iter_rows():
    if row[61].value=="Positive":
        l=[row[j].value for j in range (PET0.max_column)]
        PET2_plus.append(l)
    elif row[61].value=="Negative":
        l=[row[j].value for j in range (PET0.max_column)]
        PET2_moins.append(l)
        
        

for row in PET2_moins.iter_rows():
    if not isinstance(row[0].value ,(str)):
        l=[row[j].value for j in range(PET2_moins.max_column)]
        total_cas_spéciaux.append(l)
    
List_total_cas_speciaux = [total_cas_spéciaux.cell(j,1).value for j in range(2,total_cas_spéciaux.max_row+1)]    

for row in PET2_plus.iter_rows():
    if not isinstance(row[0].value ,(str)):
        l=[row[j].value for j in range(PET2_plus.max_column)]
        total_cas_spéciaux.append(l)  
 

List_PET2_moins = [PET2_moins.cell(j,1).value for j in range(2,PET2_moins.max_row+1)]
List_PET2_plus = [PET2_plus.cell(j,1).value for j in range(2,PET2_plus.max_row+1)]
List_PET4_plus = [PET4_plus.cell(j,1).value for j in range(2,PET4_plus.max_row+1)]
List_total_cas_speciaux = [total_cas_spéciaux.cell(j,1).value for j in range(2,total_cas_spéciaux.max_row+1)]
for row in PET4_plus.iter_rows():
    if (not isinstance(row[0].value ,(str))) and row[0].value not in List_total_cas_speciaux:
       l=[row[j].value for j in range(PET4_plus.max_column)] 
       total_cas_spéciaux.append(l)
       
List_total_cas_speciaux = [total_cas_spéciaux.cell(j,1).value for j in range(2,total_cas_spéciaux.max_row+1)]



for i in range(2,PET4_plus.max_row+1):
    if PET4_plus.cell(row=i, column=1).value in List_PET2_moins:
        l=[PET4_plus.cell(row=i, column=j).value for j in range(1, PET4_plus.max_column+1)]
        PET_2moins_4plus.append(l)
    elif PET4_plus.cell(row=i, column=1).value in List_PET2_plus:
        l=[PET4_plus.cell(row=i, column=j).value for j in range(1, PET4_plus.max_column+1)]
        PET_2plus_4plus.append(l)

wb1.save(filename1);  wb2.save(filename2);  wb3.save(filename3);   wb4.save(filename4)
wb5.save(filename5);  wb6.save(filename6);  wb7.save(filename7);   wb8.save(filename8)
wb9.save(filename9);  wb10.save(filename10);wb11.save(filename11); wb12.save(filename12)
wb13.save(filename13);wb14.save(filename14);wb15.save(filename15); wb16.save(filename16)
wb17.save(filename17)
print("END")

